#!/usr/bin/env python3
"""
GitHub Copilot Productivity Report Generator

Generates a formatted Excel workbook with per-user productivity metrics and
team-level summary KPIs using GitHub Copilot seat data and the NDJSON
users-28-day metrics report.

Usage:
    python copilot_productivity_report.py --enterprise my-ent
    python copilot_productivity_report.py --orgs org1,org2 --token ghp_xxx
    python copilot_productivity_report.py --enterprise my-ent --output-dir ./reports
"""

from __future__ import annotations

import argparse
import json
import os
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import httpx
except ImportError:
    sys.exit("ERROR: httpx is required. Install with: pip install httpx")

try:
    from dotenv import load_dotenv
except ImportError:
    load_dotenv = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl is required. Install with: pip install openpyxl")


GITHUB_API_BASE = "https://api.github.com"
API_VERSION = "2022-11-28"
REPORT_DAYS = 28
MAX_PER_PAGE = 100
MINUTES_SAVED_PER_ACCEPTANCE = 5  # conservative estimate per accepted suggestion

USER_PRODUCTIVITY_COLUMNS = [
    "organization",
    "user_login",
    "seat_assigned_date",
    "last_activity_date",
    "days_inactive",
    "active_days",
    "adoption_rate_pct",
    "total_interactions",
    "code_generations",
    "code_acceptances",
    "acceptance_rate_pct",
    "loc_suggested",
    "loc_added",
    "loc_deleted",
    "net_loc_change",
    "copilot_contribution_pct",
    "chat_interactions",
    "agent_interactions",
    "features_used",
    "engagement_depth",
    "estimated_time_saved_hrs",
    "health_profile",
    "health_notes",
]

UNIQUE_USERS_COLUMNS = [
    "organizations" if col == "organization" else col
    for col in USER_PRODUCTIVITY_COLUMNS
]

ENABLEMENT_COLUMNS = [
    "user_login",
    "organizations",
    "seat_assigned_date",
    "last_activity_date",
    "days_inactive",
    "active_days",
    "total_interactions",
    "code_generations",
    "code_acceptances",
    "health_notes",
]

HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, color="1F4E79")
LABEL_FONT = Font(bold=True)
HEALTH_STYLES = {
    "Power User": {
        "fill": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "font": Font(color="006100"),
    },
    "Healthy": {
        "fill": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "font": Font(color="006100"),
    },
    "Agent-Heavy": {
        "fill": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "font": Font(color="9C6500"),
    },
    "Chat-Focused": {
        "fill": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "font": Font(color="9C6500"),
    },
    "Moderate": {
        "fill": PatternFill(fill_type="solid", fgColor="BDD7EE"),
        "font": Font(color="003399"),
    },
    "Low Usage": {
        "fill": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "font": Font(color="9C0006"),
    },
    "Needs Enablement": {
        "fill": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "font": Font(color="9C0006"),
    },
}


def _headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": API_VERSION,
    }


def _try_print_response(resp: httpx.Response) -> None:
    try:
        body = resp.json()
        msg = body.get("message", "")
        if msg:
            print(f"    API: {msg}")
    except Exception:
        pass


def _handle_rate_limit(response: httpx.Response) -> bool:
    remaining = response.headers.get("x-ratelimit-remaining")
    if remaining is not None and remaining.isdigit() and int(remaining) == 0:
        reset_ts = int(response.headers.get("x-ratelimit-reset", 0) or 0)
        wait = max(reset_ts - int(time.time()), 1)
        print(f"    ⏳ Rate-limited. Waiting {wait}s …")
        time.sleep(wait)
        return True
    return False


def _safe_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _safe_pct(numerator: float, denominator: float, cap: float | None = None) -> float:
    if not denominator:
        return 0.0
    pct = round(numerator / denominator * 100, 1)
    if cap is not None:
        pct = min(pct, cap)
    return pct


def _parse_iso_date(value: Any) -> date | None:
    """Parse an ISO 8601 date or datetime string; return the date portion."""
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    candidate = text.replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(candidate).date()
    except ValueError:
        pass
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _format_iso_date(value: date | None) -> str:
    return value.isoformat() if value else ""


def _safe_filename(name: str) -> str:
    """Sanitize a string for safe use in a filename."""
    safe = "".join(c if c.isalnum() or c in ("-", "_") else "_" for c in (name or ""))
    return safe[:80] or "unknown"


def _compute_days_inactive(last_activity: date | None, baseline: date | None) -> Any:
    """Days between today (or baseline, whichever is later) and last activity.

    Uses today's date as the floor so the result reflects calendar reality.
    GitHub's NDJSON window typically ends 1-2 days behind the current date
    (data is batch-processed daily), so a user with activity on the last day
    of the window would otherwise show ``days_inactive = 0`` even though
    they haven't been active for a real calendar day. Returns ``'Never'``
    when no activity has ever been recorded.
    """
    if last_activity is None:
        return "Never"
    today = datetime.utcnow().date()
    if baseline is None or baseline < today:
        baseline = today
    delta = (baseline - last_activity).days
    return max(delta, 0)


def validate_token(token: str) -> None:
    with httpx.Client(timeout=30) as client:
        resp = client.get(f"{GITHUB_API_BASE}/user", headers=_headers(token))
        if resp.status_code != 200:
            sys.exit(f"ERROR: Token invalid (HTTP {resp.status_code}).")
        user = resp.json()
        scopes = resp.headers.get("x-oauth-scopes", "(unknown)")
        print(f"🔑 Authenticated as: {user.get('login', '?')}  |  Scopes: {scopes}")


def discover_orgs(token: str, enterprise: str) -> list[str]:
    """Discover orgs under an enterprise, falling back to user orgs."""
    orgs: list[str] = []
    with httpx.Client(timeout=30) as client:
        resp = client.get(
            f"{GITHUB_API_BASE}/enterprises/{enterprise}/organizations",
            headers=_headers(token),
            params={"per_page": MAX_PER_PAGE},
        )
        if resp.status_code in (403, 429) and _handle_rate_limit(resp):
            resp = client.get(
                f"{GITHUB_API_BASE}/enterprises/{enterprise}/organizations",
                headers=_headers(token),
                params={"per_page": MAX_PER_PAGE},
            )
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, list):
                orgs = [o["login"] for o in data if o.get("login")]
                if orgs:
                    return orgs

        print("  ⚠ Enterprise endpoint unavailable. Using your org memberships.")
        page = 1
        while True:
            resp = client.get(
                f"{GITHUB_API_BASE}/user/orgs",
                headers=_headers(token),
                params={"page": page, "per_page": MAX_PER_PAGE},
            )
            if resp.status_code in (403, 429) and _handle_rate_limit(resp):
                continue
            if resp.status_code != 200:
                _try_print_response(resp)
                break
            data = resp.json()
            if not data:
                break
            orgs.extend(o["login"] for o in data if o.get("login"))
            if len(data) < MAX_PER_PAGE:
                break
            page += 1
    return orgs


def fetch_seats(token: str, org: str, debug: bool = False) -> tuple[list[dict], str]:
    """Fetch Copilot seat assignments for an org.

    Returns (seats, status) where status is one of:
      - 'ok'             : seats successfully retrieved (may be empty if org has none)
      - 'forbidden'      : 403 — token lacks scope or user lacks admin permission
      - 'not_found'      : 404 — org has no Copilot subscription or endpoint not enabled
      - 'http_<code>'    : other HTTP error
      - 'error'          : transport / parse error
    """
    seats: list[dict] = []
    page = 1
    last_status = "ok"
    raw_pages: list[dict] = []
    with httpx.Client(timeout=30) as client:
        while True:
            try:
                resp = client.get(
                    f"{GITHUB_API_BASE}/orgs/{org}/copilot/billing/seats",
                    headers=_headers(token),
                    params={"page": page, "per_page": MAX_PER_PAGE},
                )
            except (httpx.ConnectError, httpx.TimeoutException) as exc:
                print(f"   ⚠ Network error fetching seats for {org}: {exc}")
                return seats, "error"

            if resp.status_code in (403, 429) and _handle_rate_limit(resp):
                continue
            if resp.status_code != 200:
                if resp.status_code == 403:
                    last_status = "forbidden"
                    print(
                        f"   ⚠ WARNING: seats fetch returned HTTP 403 for org '{org}'. "
                        f"Token likely lacks 'manage_billing:copilot' scope OR you are not an admin of this org. "
                        f"Seat dates (assigned/last-used) will be blank for users in this org."
                    )
                elif resp.status_code == 404:
                    last_status = "not_found"
                    print(
                        f"   ⚠ WARNING: seats endpoint returned HTTP 404 for org '{org}'. "
                        f"Org may not have a Copilot Business/Enterprise subscription. "
                        f"Seat dates will be blank for users in this org."
                    )
                else:
                    last_status = f"http_{resp.status_code}"
                    print(
                        f"   ⚠ WARNING: seats fetch returned HTTP {resp.status_code} for org '{org}'. "
                        f"Seat dates will be blank for users in this org."
                    )
                _try_print_response(resp)
                break
            try:
                data = resp.json()
            except json.JSONDecodeError:
                last_status = "error"
                print(f"   ⚠ WARNING: seats response for org '{org}' was not valid JSON.")
                break
            raw_pages.append(data)
            page_seats = data.get("seats", [])
            if not page_seats:
                break
            seats.extend(page_seats)
            if len(seats) >= data.get("total_seats", len(seats)):
                break
            page += 1

    if debug and (raw_pages or last_status != "ok"):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        debug_path = Path(f"debug_seats_{_safe_filename(org)}_{ts}.json")
        try:
            debug_path.write_text(
                json.dumps({"status": last_status, "pages": raw_pages}, indent=2),
                encoding="utf-8",
            )
            print(f"   🐛 [debug] saved raw seats payload → {debug_path}")
            if seats:
                first_keys = sorted(seats[0].keys())
                print(f"   🐛 [debug] first seat top-level keys: {first_keys}")
                assignee = seats[0].get("assignee") or {}
                if isinstance(assignee, dict):
                    print(f"   🐛 [debug] first seat assignee keys: {sorted(assignee.keys())}")
        except OSError as exc:
            print(f"   ⚠ Could not write debug seats payload: {exc}")

    return seats, last_status


def fetch_enterprise_seats(
    token: str, enterprise: str, debug: bool = False
) -> tuple[dict[str, list[dict]], str]:
    """Fetch Copilot seats from the ENTERPRISE-level endpoint and group by org.

    Returns (org_to_seats, status). Used as a fallback when the per-org
    endpoint returns empty for orgs managed under an enterprise-wide
    Copilot subscription. Each seat in the response carries an
    ``organization`` field telling us which child org granted the seat.

    Endpoint:  GET /enterprises/{enterprise}/copilot/billing/seats
    Required:  manage_billing:copilot  OR  read:enterprise scope, AND the
               user must be an enterprise owner or billing manager.
    """
    org_to_seats: dict[str, list[dict]] = {}
    page = 1
    last_status = "ok"
    raw_pages: list[dict] = []
    total_seats = 0
    with httpx.Client(timeout=30) as client:
        while True:
            try:
                resp = client.get(
                    f"{GITHUB_API_BASE}/enterprises/{enterprise}/copilot/billing/seats",
                    headers=_headers(token),
                    params={"page": page, "per_page": MAX_PER_PAGE},
                )
            except (httpx.ConnectError, httpx.TimeoutException) as exc:
                print(f"   ⚠ Network error fetching enterprise seats: {exc}")
                return org_to_seats, "error"

            if resp.status_code in (403, 429) and _handle_rate_limit(resp):
                continue
            if resp.status_code != 200:
                if resp.status_code == 403:
                    last_status = "forbidden"
                    print(
                        f"   ⚠ Enterprise seats endpoint returned HTTP 403. "
                        f"Token needs 'manage_billing:copilot' or 'read:enterprise', "
                        f"AND you must be an enterprise owner / billing manager."
                    )
                elif resp.status_code == 404:
                    last_status = "not_found"
                    print(
                        f"   ⚠ Enterprise seats endpoint returned HTTP 404. "
                        f"Enterprise '{enterprise}' either doesn't exist or has no "
                        f"Copilot Business/Enterprise subscription."
                    )
                else:
                    last_status = f"http_{resp.status_code}"
                    print(
                        f"   ⚠ Enterprise seats endpoint returned HTTP "
                        f"{resp.status_code}."
                    )
                _try_print_response(resp)
                break
            try:
                data = resp.json()
            except json.JSONDecodeError:
                last_status = "error"
                print("   ⚠ Enterprise seats response was not valid JSON.")
                break
            raw_pages.append(data)
            page_seats = data.get("seats", []) or []
            if not page_seats:
                break
            for seat in page_seats:
                org_obj = seat.get("organization") or {}
                org_login = org_obj.get("login") if isinstance(org_obj, dict) else None
                if not org_login:
                    # Some enterprise seats are granted via enterprise teams
                    # rather than orgs. Bucket those under a synthetic key so
                    # we don't drop the seat data entirely.
                    org_login = "__enterprise_team__"
                org_to_seats.setdefault(org_login, []).append(seat)
                total_seats += 1
            if total_seats >= data.get("total_seats", total_seats):
                break
            page += 1

    if debug and (raw_pages or last_status != "ok"):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        debug_path = Path(
            f"debug_enterprise_seats_{_safe_filename(enterprise)}_{ts}.json"
        )
        try:
            debug_path.write_text(
                json.dumps(
                    {"status": last_status, "pages": raw_pages}, indent=2
                ),
                encoding="utf-8",
            )
            print(f"   🐛 [debug] saved raw enterprise seats payload → {debug_path}")
        except OSError as exc:
            print(f"   ⚠ Could not write debug enterprise seats payload: {exc}")

    return org_to_seats, last_status


def fetch_ndjson(token: str, org: str, debug: bool = False) -> tuple[list[dict], str, str]:
    """Fetch per-user NDJSON usage metrics and report period metadata."""
    records: list[dict] = []
    report_start = ""
    report_end = ""
    raw_text_chunks: list[str] = []
    with httpx.Client(timeout=60) as client:
        while True:
            resp = client.get(
                f"{GITHUB_API_BASE}/orgs/{org}/copilot/metrics/reports/users-28-day/latest",
                headers=_headers(token),
            )
            if resp.status_code in (403, 429) and _handle_rate_limit(resp):
                continue
            if resp.status_code != 200:
                if debug:
                    print(
                        f"   🐛 [debug] NDJSON manifest fetch for {org} returned HTTP {resp.status_code}"
                    )
                    _try_print_response(resp)
                return [], report_start, report_end
            break

        payload = resp.json()
        report_start = payload.get("report_start_day", "") or ""
        report_end = payload.get("report_end_day", "") or ""
        links = payload.get("download_links", [])

        if links:
            print(f"   📥 NDJSON report: {report_start} → {report_end} ({len(links)} file(s))")

        for link in links:
            try:
                dl = client.get(link, timeout=120)
                if dl.status_code in (401, 403):
                    dl = client.get(link, headers=_headers(token), timeout=120)
            except httpx.ConnectError:
                print(f"   ⚠ Network error downloading NDJSON file — retrying …")
                try:
                    import time as _time
                    _time.sleep(3)
                    dl = client.get(link, headers=_headers(token), timeout=120)
                except httpx.ConnectError:
                    print(f"   ❌ Download failed (DNS/network error). Check your internet connection.")
                    continue
            except httpx.TimeoutException:
                print(f"   ⚠ Download timed out — skipping file.")
                continue
            if dl.status_code != 200:
                continue
            if debug:
                raw_text_chunks.append(dl.text)
            for line in dl.text.strip().splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    records.append(json.loads(line))
                except json.JSONDecodeError:
                    continue

    if debug:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        if raw_text_chunks:
            debug_path = Path(f"debug_{_safe_filename(org)}_{ts}.ndjson")
            try:
                debug_path.write_text("\n".join(raw_text_chunks), encoding="utf-8")
                print(f"   🐛 [debug] saved raw NDJSON → {debug_path}")
            except OSError as exc:
                print(f"   ⚠ Could not write debug NDJSON: {exc}")
        if records:
            first = records[0]
            print(f"   🐛 [debug] first NDJSON record has {len(first)} fields")
            print(f"   🐛 [debug] first record top-level keys: {sorted(first.keys())}")
            # Highlight chat/agent specific fields so the user can spot mismatches.
            relevant = {
                k: v for k, v in first.items()
                if "chat" in k.lower() or "agent" in k.lower() or "panel" in k.lower()
            }
            if relevant:
                print(f"   🐛 [debug] chat/agent-related fields on first record: {relevant}")
            else:
                print("   🐛 [debug] no chat/agent/panel fields found on first record")

    return records, report_start, report_end


def dedupe_ndjson_records(records: list[dict]) -> tuple[list[dict], int]:
    """Dedupe NDJSON records by (user_login, day).

    GitHub's user-level NDJSON returns each user's GLOBAL Copilot activity for
    every org they hold a seat in — not org-scoped activity. So if a user is in
    N orgs we will see N identical records per active day. Dedupe by
    (user_login, day).

    Defensive merge: in normal operation duplicates are byte-for-byte
    identical, but to guard against any silent divergence we merge field-wise
    — taking the MAX of numeric/count fields and the OR of boolean flags —
    so we never lose volume data. Differences are logged.

    Returns (merged_records, removed_count).
    """
    merged: dict[tuple[str, str], dict] = {}
    diff_keys: set[str] = set()
    diff_pairs = 0
    for rec in records:
        login = rec.get("user_login") or ""
        day = rec.get("day") or ""
        if not login or not day:
            continue
        key = (login, day)
        existing = merged.get(key)
        if existing is None:
            merged[key] = dict(rec)
            continue

        pair_had_diff = False
        for field, new_val in rec.items():
            old_val = existing.get(field)
            if isinstance(new_val, bool) or isinstance(old_val, bool):
                combined = bool(new_val) or bool(old_val)
                if combined != bool(old_val):
                    existing[field] = combined
                    pair_had_diff = True
                    diff_keys.add(field)
            elif isinstance(new_val, (int, float)) or isinstance(old_val, (int, float)):
                try:
                    new_num = float(new_val or 0)
                    old_num = float(old_val or 0)
                except (TypeError, ValueError):
                    continue
                if new_num > old_num:
                    existing[field] = new_val
                    pair_had_diff = True
                    diff_keys.add(field)
                elif new_num < old_num:
                    pair_had_diff = True
                    diff_keys.add(field)
            elif old_val is None and new_val is not None:
                existing[field] = new_val
        if pair_had_diff:
            diff_pairs += 1

    removed = len(records) - len(merged)
    if diff_pairs:
        print(
            f"   ⚠ Note: {diff_pairs} (user, day) pair(s) had differing values across orgs "
            f"on fields: {sorted(diff_keys)}. Merged by field-wise max / boolean OR so no "
            f"volume data was lost. (Normally records should be identical across orgs — this "
            f"may indicate GitHub returned stale data for one org.)"
        )
    return list(merged.values()), removed


def aggregate_user_ndjson(records: list[dict]) -> dict[str, dict[str, Any]]:
    """Aggregate daily NDJSON records per user_login."""
    users: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "days_seen": set(),
            "total_interactions": 0,
            "code_generations": 0,
            "code_acceptances": 0,
            "loc_suggested": 0,
            "loc_added": 0,
            "loc_deleted": 0,
            "chat_interactions": 0,
            "agent_interactions": 0,
            "used_chat": False,
            "used_agent": False,
            "used_cli": False,
            "used_code_review": False,
        }
    )

    for rec in records:
        login = rec.get("user_login", "")
        if not login:
            continue

        user = users[login]
        day = rec.get("day")
        if day:
            user["days_seen"].add(day)

        user["total_interactions"] += _safe_int(rec.get("user_initiated_interaction_count"))
        user["code_generations"] += _safe_int(rec.get("code_generation_activity_count"))
        user["code_acceptances"] += _safe_int(rec.get("code_acceptance_activity_count"))
        user["loc_suggested"] += _safe_int(rec.get("loc_suggested_to_add_sum"))
        user["loc_added"] += _safe_int(rec.get("loc_added_sum"))
        user["loc_deleted"] += _safe_int(rec.get("loc_deleted_sum"))
        user["chat_interactions"] += (
            _safe_int(rec.get("chat_panel_ask_mode"))
            + _safe_int(rec.get("chat_panel_edit_mode"))
            + _safe_int(rec.get("chat_panel_plan_mode"))
            + _safe_int(rec.get("chat_panel_custom_mode"))
        )
        user["agent_interactions"] += _safe_int(rec.get("chat_panel_agent_mode"))

        if rec.get("used_chat"):
            user["used_chat"] = True
        if rec.get("used_agent"):
            user["used_agent"] = True
        if rec.get("used_cli"):
            user["used_cli"] = True
        if rec.get("used_copilot_code_review_active") or rec.get("used_copilot_code_review_passive"):
            user["used_code_review"] = True

    aggregated: dict[str, dict[str, Any]] = {}
    for login, metrics in users.items():
        aggregated[login] = {
            **metrics,
            "active_days": len(metrics["days_seen"]),
            "active_day_set": frozenset(metrics["days_seen"]),
        }
        aggregated[login].pop("days_seen", None)
    return aggregated


def build_features_list(metrics: dict[str, Any]) -> str:
    features: list[str] = []
    if metrics.get("used_chat"):
        features.append("chat")
    if metrics.get("used_agent"):
        features.append("agent")
    if metrics.get("used_cli"):
        features.append("cli")
    if metrics.get("used_code_review"):
        features.append("code_review")
    return ", ".join(features) if features else "none"


def classify_health(metrics: dict[str, Any]) -> tuple[str, str]:
    active_days = metrics["active_days"]
    total_interactions = metrics["total_interactions"]
    acceptance_rate = metrics["acceptance_rate_pct"]
    engagement_depth = metrics["engagement_depth"]
    chat_interactions = metrics["chat_interactions"]
    agent_interactions = metrics["agent_interactions"]
    loc_added = metrics["loc_added"]
    loc_suggested = metrics["loc_suggested"]
    code_generations = metrics["code_generations"]

    if active_days == 0 or (active_days > 0 and total_interactions < 5):
        return "Needs Enablement", "No meaningful usage detected — training recommended"
    if 1 <= active_days <= 3 and total_interactions < 20:
        return "Low Usage", "Minimal usage — may need guidance on workflow integration"
    if acceptance_rate >= 30 and active_days >= 14 and engagement_depth >= 50:
        return "Power User", "Heavy, effective usage across features"
    if agent_interactions > chat_interactions and loc_added > loc_suggested * 2 and loc_added > 0:
        return "Agent-Heavy", "Primary usage through agent/edit mode — advanced adoption pattern"
    if chat_interactions > 0 and code_generations <= 5:
        return "Chat-Focused", "Uses Copilot for Q&A and understanding rather than code generation"
    if acceptance_rate >= 25 and active_days >= 7:
        return "Healthy", "Good adoption with consistent usage"
    return "Moderate", "Active but room to increase engagement"


def build_user_rows(
    org: str,
    seats: list[dict],
    ndjson_records: list[dict] | None = None,
    report_end: str = "",
    global_aggregation: dict[str, dict[str, Any]] | None = None,
    global_seat_map: dict[str, dict] | None = None,
    debug: bool = False,
) -> list[dict[str, Any]]:
    """Build per-user rows for one org.

    If `global_aggregation` is provided (a dict from user_login → aggregated
    metrics produced by `aggregate_user_ndjson()` over the GLOBAL deduped
    record pool), per-user numbers come from there. This is the correct path
    because the user-level NDJSON returns global per-user activity for every
    org the user holds a seat in — aggregating per-org would 4× overcount a
    user who's in 4 orgs.

    If `global_aggregation` is None, falls back to aggregating the supplied
    org-scoped `ndjson_records` (legacy / single-org callers like the mock
    generator).

    `global_seat_map` (login.lower() → seat dict) is consulted as a fallback
    for users that don't have a per-org seat — typically because the user
    holds a seat granted by a DIFFERENT org under the same enterprise
    (Copilot Enterprise spillover), or because the user is licensed via
    Copilot Pro showing up in another org's activity. This ensures
    `seat_assigned_date`, `last_activity_date`, and `plan_type` populate
    for every licensed user across every org row they appear in.
    """
    # Case-insensitive seat lookup: NDJSON typically returns lowercase logins,
    # but the seats endpoint preserves the original GitHub casing. Without case
    # normalization, a user `Pthangavel_Pebin` in seats would never match
    # `pthangavel_pebin` in NDJSON — seat dates would silently drop to None
    # AND the user could appear as two separate rows (one from seats, one from
    # NDJSON). We key all lookups by `login.lower()` and prefer the NDJSON
    # casing for the display login when both sources have the user.
    seat_map: dict[str, dict] = {}
    seat_display_login: dict[str, str] = {}
    seats_with_login = 0
    seats_with_created_at = 0
    for seat in seats:
        assignee = seat.get("assignee", {}) or {}
        login = assignee.get("login", "")
        if login:
            key = login.lower()
            seat_map[key] = seat
            seat_display_login[key] = login
            seats_with_login += 1
            if seat.get("created_at"):
                seats_with_created_at += 1

    if debug and seats:
        first_seat = seats[0]
        first_keys = sorted(first_seat.keys())
        assignee_keys = sorted((first_seat.get("assignee") or {}).keys()) if isinstance(first_seat.get("assignee"), dict) else []
        print(
            f"   🐛 [debug] {org}: {len(seats)} seat(s) fetched, "
            f"{seats_with_login} have assignee.login, "
            f"{seats_with_created_at} have non-null created_at"
        )
        print(f"   🐛 [debug] first seat top-level keys:  {first_keys}")
        print(f"   🐛 [debug] first seat assignee.* keys: {assignee_keys}")
        if seats_with_created_at == 0:
            print(
                f"   ⚠ All seats in {org} have a null/missing 'created_at' field. "
                f"GitHub may have renamed it. Inspect the saved debug_seats_*.json."
            )

    baseline_date = _parse_iso_date(report_end) or datetime.utcnow().date()
    if global_aggregation is not None:
        ndjson_agg = global_aggregation
        org_logins = {
            rec.get("user_login")
            for rec in (ndjson_records or [])
            if rec.get("user_login")
        }
    else:
        ndjson_agg = aggregate_user_ndjson(ndjson_records or [])
        org_logins = set(ndjson_agg.keys())

    # Union of all logins (case-folded). Prefer NDJSON's casing for display
    # because that's what GitHub uses in audit logs / activity dashboards;
    # fall back to the seat assignee.login casing for seat-only users.
    canonical: dict[str, str] = {}
    for login in org_logins:
        if login:
            canonical[login.lower()] = login
    for key, display in seat_display_login.items():
        canonical.setdefault(key, display)

    all_keys = sorted(canonical.keys())
    rows: list[dict[str, Any]] = []

    for key in all_keys:
        login = canonical[key]
        aggregated = ndjson_agg.get(login) or ndjson_agg.get(key) or {}
        active_days = aggregated.get("active_days", 0)
        total_interactions = aggregated.get("total_interactions", 0)
        code_generations = aggregated.get("code_generations", 0)
        code_acceptances = aggregated.get("code_acceptances", 0)
        loc_suggested = aggregated.get("loc_suggested", 0)
        loc_added = aggregated.get("loc_added", 0)
        loc_deleted = aggregated.get("loc_deleted", 0)
        chat_interactions = aggregated.get("chat_interactions", 0)
        agent_interactions = aggregated.get("agent_interactions", 0)
        # Engagement depth = active prompts sent to Copilot across ALL surfaces
        # (chat panel, inline chat, terminal chat, github.com chat, agent). This
        # is `user_initiated_interaction_count` which by definition excludes passive
        # completions, mode-switches, and shortcuts. Using only the chat_panel_*_mode
        # counters undercounted users who use chat outside the panel (e.g. inline chat
        # in the editor) — they'd show engagement_depth=0 despite used_chat=true.
        # chat_interactions and agent_interactions are kept as breakdown columns and
        # are populated only when chat_panel_*_mode counters are populated.
        engagement_depth = total_interactions

        seat = seat_map.get(key, {})
        seat_source = "per_org" if seat else None
        # Fallback: a licensed user may not have an org-assigned seat in this
        # specific org (e.g. their seat was granted by another org under the
        # same enterprise, or they're on Copilot Pro). Pull seat data from
        # the global seat map keyed by `login.lower()`. This is what makes
        # `seat_assigned_date` populate for users whose enterprise license
        # was granted by an org we didn't iterate as the "current" org.
        if not seat and global_seat_map:
            ent_seat = global_seat_map.get(key)
            if ent_seat:
                seat = ent_seat
                seat_source = "enterprise"

        assigned_date = _parse_iso_date(seat.get("created_at"))
        last_activity_date = _parse_iso_date(seat.get("last_activity_at"))

        # NDJSON-backfill for last_activity_date: when seats are unavailable
        # (token scope / admin role / 404 etc.) we can still derive the last
        # activity date from the user's max active day in NDJSON. This won't
        # help inactive licensed users (they don't appear in NDJSON), but for
        # active users it means the column is populated and days_inactive is
        # computable even when the seats API is silently failing.
        last_activity_source = "seats"
        if last_activity_date is None:
            day_set = aggregated.get("active_day_set") or frozenset()
            if day_set:
                last_activity_date = max(_parse_iso_date(d) or date.min for d in day_set)
                if last_activity_date == date.min:
                    last_activity_date = None
                else:
                    last_activity_source = "ndjson"

        days_inactive = _compute_days_inactive(last_activity_date, baseline_date)

        row: dict[str, Any] = {
            "organization": org,
            "user_login": login,
            "seat_assigned_date": _format_iso_date(assigned_date),
            "last_activity_date": _format_iso_date(last_activity_date),
            "days_inactive": days_inactive,
            "active_days": active_days,
            "adoption_rate_pct": round(active_days / REPORT_DAYS * 100, 1),
            "total_interactions": total_interactions,
            "code_generations": code_generations,
            "code_acceptances": code_acceptances,
            "acceptance_rate_pct": _safe_pct(code_acceptances, code_generations),
            "loc_suggested": loc_suggested,
            "loc_added": loc_added,
            "loc_deleted": loc_deleted,
            "net_loc_change": loc_added - loc_deleted,
            "copilot_contribution_pct": _safe_pct(loc_suggested, loc_added, cap=100.0),
            "chat_interactions": chat_interactions,
            "agent_interactions": agent_interactions,
            "features_used": build_features_list(aggregated),
            "engagement_depth": engagement_depth,
            "estimated_time_saved_hrs": round(code_acceptances * MINUTES_SAVED_PER_ACCEPTANCE / 60, 1),
            "used_chat": bool(aggregated.get("used_chat", False)),
            "used_agent": bool(aggregated.get("used_agent", False)),
            "used_cli": bool(aggregated.get("used_cli", False)),
            "used_code_review": bool(aggregated.get("used_code_review", False)),
            "plan_type": (seat.get("plan_type") or "") if seat else "",
            "_active_day_set": aggregated.get("active_day_set", frozenset()),
        }
        row["health_profile"], row["health_notes"] = classify_health(row)
        rows.append(row)

    rows.sort(key=lambda item: (item["organization"], item["user_login"].lower()))
    return rows


def dedupe_users_across_orgs(
    user_rows: list[dict[str, Any]],
    report_end: str = "",
    metrics_are_global: bool = False,
) -> list[dict[str, Any]]:
    """Collapse per-org rows into one row per user_login.

    organizations is a comma-separated, sorted list. Seat dates use earliest
    assigned and latest activity; days_inactive is recomputed against the
    shared baseline (report_end). Health is reclassified on the merged numbers.

    When ``metrics_are_global=True`` (the live `main()` path), each per-org row
    already contains the user's GLOBAL Copilot metrics (because the user-level
    NDJSON returns global activity for every org the user is a seat in). In
    that case we take the first row's metrics directly — summing would multiply
    by the number of orgs the user is in. Otherwise (mock / legacy / truly
    org-scoped data) volume metrics are summed and `active_days` uses the
    union of distinct days across orgs (capped at 28).
    """
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in user_rows:
        grouped[row["user_login"]].append(row)

    baseline_date = _parse_iso_date(report_end) or datetime.utcnow().date()
    merged_rows: list[dict[str, Any]] = []

    sum_fields = (
        "total_interactions",
        "code_generations",
        "code_acceptances",
        "loc_suggested",
        "loc_added",
        "loc_deleted",
        "chat_interactions",
        "agent_interactions",
    )

    for login, group in grouped.items():
        organizations = sorted({row["organization"] for row in group if row.get("organization")})

        merged: dict[str, Any] = {
            "organizations": ", ".join(organizations),
            "user_login": login,
        }

        if metrics_are_global:
            # Each per-org row already has the user's global metrics — take one.
            base = group[0]
            for field in sum_fields:
                merged[field] = base.get(field, 0)
            merged["active_days"] = base.get("active_days", 0)
        else:
            for field in sum_fields:
                merged[field] = sum(row.get(field, 0) for row in group)
            union_days: set[str] = set()
            for row in group:
                union_days |= set(row.get("_active_day_set") or ())
            merged["active_days"] = min(len(union_days), REPORT_DAYS) if union_days else max(
                (row.get("active_days", 0) for row in group), default=0
            )

        merged["adoption_rate_pct"] = round(merged["active_days"] / REPORT_DAYS * 100, 1)
        merged["acceptance_rate_pct"] = _safe_pct(merged["code_acceptances"], merged["code_generations"])
        merged["net_loc_change"] = merged["loc_added"] - merged["loc_deleted"]
        merged["copilot_contribution_pct"] = _safe_pct(merged["loc_suggested"], merged["loc_added"], cap=100.0)
        merged["engagement_depth"] = merged["total_interactions"]
        merged["estimated_time_saved_hrs"] = round(
            merged["code_acceptances"] * MINUTES_SAVED_PER_ACCEPTANCE / 60, 1
        )

        merged["used_chat"] = any(row.get("used_chat") for row in group)
        merged["used_agent"] = any(row.get("used_agent") for row in group)
        merged["used_cli"] = any(row.get("used_cli") for row in group)
        merged["used_code_review"] = any(row.get("used_code_review") for row in group)
        merged["features_used"] = build_features_list(merged)

        assigned_dates = [d for d in (_parse_iso_date(row.get("seat_assigned_date")) for row in group) if d]
        activity_dates = [d for d in (_parse_iso_date(row.get("last_activity_date")) for row in group) if d]
        merged["seat_assigned_date"] = _format_iso_date(min(assigned_dates)) if assigned_dates else ""
        merged["last_activity_date"] = _format_iso_date(max(activity_dates)) if activity_dates else ""
        merged["days_inactive"] = _compute_days_inactive(
            max(activity_dates) if activity_dates else None, baseline_date
        )

        plan_types = sorted({row.get("plan_type") for row in group if row.get("plan_type")})
        merged["plan_type"] = ", ".join(plan_types)

        merged["health_profile"], merged["health_notes"] = classify_health(merged)
        merged_rows.append(merged)

    merged_rows.sort(key=lambda item: item["user_login"].lower())
    return merged_rows


def build_team_summary(
    user_rows: list[dict[str, Any]],
    report_start: str,
    report_end: str,
    unique_rows: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    if unique_rows is None:
        unique_rows = dedupe_users_across_orgs(user_rows, report_end=report_end)

    total_users = len(user_rows)
    active_users = [row for row in user_rows if row["active_days"] > 0]
    active_user_count = len(active_users)

    unique_user_count = len(unique_rows)
    unique_active_user_count = sum(1 for row in unique_rows if row.get("active_days", 0) > 0)

    # IMPORTANT: totals are summed over unique users, not per-org rows. The user-level
    # NDJSON returns each user's GLOBAL metrics for every org they hold a seat in, so
    # summing per-org rows would multiply totals by the number of orgs per user.
    total_code_generations = sum(row["code_generations"] for row in unique_rows)
    total_code_acceptances = sum(row["code_acceptances"] for row in unique_rows)
    total_loc_suggested = sum(row["loc_suggested"] for row in unique_rows)
    total_loc_added = sum(row["loc_added"] for row in unique_rows)
    total_loc_deleted = sum(row["loc_deleted"] for row in unique_rows)
    total_interactions = sum(row["total_interactions"] for row in unique_rows)
    total_estimated_time_saved = sum(row["estimated_time_saved_hrs"] for row in unique_rows)
    avg_time_saved_per_active_user = round(
        total_estimated_time_saved / unique_active_user_count, 1
    ) if unique_active_user_count else 0.0

    health_order = [
        "Power User",
        "Healthy",
        "Agent-Heavy",
        "Chat-Focused",
        "Moderate",
        "Low Usage",
        "Needs Enablement",
    ]
    distribution = Counter(row["health_profile"] for row in unique_rows)

    top_users = sorted(
        (row for row in unique_rows if row.get("engagement_depth", 0) > 0),
        key=lambda row: (
            row["engagement_depth"],
            row["active_days"],
            row["total_interactions"],
            row["user_login"].lower(),
        ),
        reverse=True,
    )[:10]
    enablement_rows = [row for row in unique_rows if row.get("health_profile") == "Needs Enablement"]

    def feature_pct(flag: str) -> float:
        if not unique_active_user_count:
            return 0.0
        count = sum(1 for row in unique_rows if row.get("active_days", 0) > 0 and row.get(flag))
        return round(count / unique_active_user_count * 100, 1)

    return {
        "total_users": total_users,
        "active_users": active_user_count,
        "unique_user_count": unique_user_count,
        "unique_active_user_count": unique_active_user_count,
        "team_adoption_rate_pct": _safe_pct(unique_active_user_count, unique_user_count),
        "report_period": f"{report_start} → {report_end}" if report_start and report_end else "N/A",
        "total_code_generations": total_code_generations,
        "total_code_acceptances": total_code_acceptances,
        "team_acceptance_rate_pct": _safe_pct(total_code_acceptances, total_code_generations),
        "total_loc_suggested": total_loc_suggested,
        "total_loc_added": total_loc_added,
        "total_loc_deleted": total_loc_deleted,
        "team_copilot_contribution_pct": _safe_pct(total_loc_suggested, total_loc_added, cap=100.0),
        "total_estimated_time_saved_hrs": round(total_estimated_time_saved, 1),
        "avg_time_saved_per_active_user_hrs": avg_time_saved_per_active_user,
        "total_interactions": total_interactions,
        "avg_engagement_depth_per_user": round(
            sum(row["engagement_depth"] for row in unique_rows) / unique_user_count, 1
        ) if unique_user_count else 0.0,
        "avg_active_days_per_user": round(
            sum(row["active_days"] for row in unique_rows) / unique_user_count, 1
        ) if unique_user_count else 0.0,
        "feature_adoption": {
            "Chat": feature_pct("used_chat"),
            "Agent": feature_pct("used_agent"),
            "CLI": feature_pct("used_cli"),
            "Code Review": feature_pct("used_code_review"),
        },
        "health_distribution": {label: distribution.get(label, 0) for label in health_order},
        "top_users": top_users,
        "enablement_users": [row["user_login"] for row in enablement_rows],
        "enablement_rows": enablement_rows,
        "unique_rows": unique_rows,
    }


def _style_header_row(ws) -> None:
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 30)


def _write_user_sheet(
    workbook: Workbook,
    sheet_title: str,
    columns: list[str],
    rows: list[dict[str, Any]],
) -> None:
    """Write a User Productivity-style sheet (per-org or unique-user variant)."""
    ws = workbook.create_sheet(sheet_title)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(column, "") for column in columns])

    _style_header_row(ws)

    percentage_columns = {"adoption_rate_pct", "acceptance_rate_pct", "copilot_contribution_pct"}
    decimal_columns = {"estimated_time_saved_hrs"}
    number_columns = {
        "active_days",
        "total_interactions",
        "code_generations",
        "code_acceptances",
        "loc_suggested",
        "loc_added",
        "loc_deleted",
        "net_loc_change",
        "chat_interactions",
        "agent_interactions",
        "engagement_depth",
    }
    date_columns = {"seat_assigned_date", "last_activity_date"}
    days_inactive_column = "days_inactive"

    header_index = {name: idx + 1 for idx, name in enumerate(columns)}
    health_col = header_index.get("health_profile")

    for row_idx in range(2, ws.max_row + 1):
        for column_name in percentage_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "0.0"
        for column_name in decimal_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "0.0"
        for column_name in number_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "#,##0"
        for column_name in date_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "yyyy-mm-dd"
        if days_inactive_column in header_index:
            cell = ws.cell(row=row_idx, column=header_index[days_inactive_column])
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"

        if health_col:
            health_cell = ws.cell(row=row_idx, column=health_col)
            style = HEALTH_STYLES.get(str(health_cell.value), HEALTH_STYLES["Moderate"])
            health_cell.fill = style["fill"]
            health_cell.font = style["font"]

    _autosize_columns(ws)


def _enablement_sort_key(row: dict[str, Any]) -> tuple[int, int, str]:
    """Sort key for Needs Enablement: 'Never' first, then largest days_inactive."""
    days = row.get("days_inactive")
    if isinstance(days, str):  # 'Never'
        return (0, 0, row["user_login"].lower())
    return (1, -int(days or 0), row["user_login"].lower())


def _write_enablement_sheet(
    workbook: Workbook,
    enablement_rows: list[dict[str, Any]],
) -> None:
    ws = workbook.create_sheet("Needs Enablement")
    ws.append(ENABLEMENT_COLUMNS)

    if not enablement_rows:
        ws.append(["No users currently flagged for enablement"] + [""] * (len(ENABLEMENT_COLUMNS) - 1))
        _style_header_row(ws)
        _autosize_columns(ws)
        return

    sorted_rows = sorted(enablement_rows, key=_enablement_sort_key)
    for row in sorted_rows:
        ws.append([row.get(column, "") for column in ENABLEMENT_COLUMNS])

    _style_header_row(ws)

    header_index = {name: idx + 1 for idx, name in enumerate(ENABLEMENT_COLUMNS)}
    number_columns = {"active_days", "total_interactions", "code_generations", "code_acceptances"}
    date_columns = {"seat_assigned_date", "last_activity_date"}

    enablement_fill = HEALTH_STYLES["Needs Enablement"]["fill"]
    enablement_font = HEALTH_STYLES["Needs Enablement"]["font"]
    login_col = header_index["user_login"]

    for row_idx in range(2, ws.max_row + 1):
        for column_name in number_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "#,##0"
        for column_name in date_columns:
            if column_name in header_index:
                ws.cell(row=row_idx, column=header_index[column_name]).number_format = "yyyy-mm-dd"
        days_cell = ws.cell(row=row_idx, column=header_index["days_inactive"])
        if isinstance(days_cell.value, (int, float)):
            days_cell.number_format = "#,##0"

        login_cell = ws.cell(row=row_idx, column=login_col)
        login_cell.fill = enablement_fill
        login_cell.font = enablement_font

    _autosize_columns(ws)


def write_excel(output_path: Path, user_rows: list[dict[str, Any]], summary: dict[str, Any]) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_user_sheet(workbook, "User Productivity", USER_PRODUCTIVITY_COLUMNS, user_rows)

    unique_rows = summary["unique_rows"]
    unique_rows_sorted = sorted(
        unique_rows,
        key=lambda row: (
            -int(row.get("engagement_depth", 0) or 0),
            -int(row.get("active_days", 0) or 0),
            row["user_login"].lower(),
        ),
    )
    _write_user_sheet(workbook, "Unique Users", UNIQUE_USERS_COLUMNS, unique_rows_sorted)

    _write_enablement_sheet(workbook, summary["enablement_rows"])

    summary_ws = workbook.create_sheet("Team Summary")
    summary_rows: list[tuple[str, Any, str]] = [
        ("Overview", "", "section"),
        ("Total User-Org Rows", summary["total_users"], "number"),
        ("Unique Users", summary["unique_user_count"], "number"),
        ("Unique Active Users", summary["unique_active_user_count"], "number"),
        ("Team Adoption Rate", summary["team_adoption_rate_pct"], "pct"),
        ("Report Period", summary["report_period"], "label"),
        ("", "", "blank"),
        ("Code Acceleration", "", "section"),
        ("Total Code Generations", summary["total_code_generations"], "number"),
        ("Total Code Acceptances", summary["total_code_acceptances"], "number"),
        ("Team Acceptance Rate", summary["team_acceptance_rate_pct"], "pct"),
        ("Total LOC Suggested", summary["total_loc_suggested"], "number"),
        ("Total LOC Added", summary["total_loc_added"], "number"),
        ("Total LOC Deleted", summary["total_loc_deleted"], "number"),
        ("Team Copilot Contribution %", summary["team_copilot_contribution_pct"], "pct"),
        ("", "", "blank"),
        ("Productivity Impact", "", "section"),
        ("Total Estimated Time Saved (hrs)", summary["total_estimated_time_saved_hrs"], "pct"),
        ("Avg Time Saved per Active User (hrs)", summary["avg_time_saved_per_active_user_hrs"], "pct"),
        ("Estimation Basis", f"{MINUTES_SAVED_PER_ACCEPTANCE} min per accepted suggestion (conservative)", "label"),
        ("", "", "blank"),
        ("Engagement", "", "section"),
        ("Total Interactions", summary["total_interactions"], "number"),
        ("Avg Engagement Depth per User", summary["avg_engagement_depth_per_user"], "pct"),
        ("Avg Active Days per User", summary["avg_active_days_per_user"], "pct"),
        ("", "", "blank"),
        ("Feature Adoption (% of active users)", "", "section"),
        ("Chat", summary["feature_adoption"]["Chat"], "pct"),
        ("Agent", summary["feature_adoption"]["Agent"], "pct"),
        ("CLI", summary["feature_adoption"]["CLI"], "pct"),
        ("Code Review", summary["feature_adoption"]["Code Review"], "pct"),
        ("", "", "blank"),
        ("Health Distribution (unique users)", "", "section"),
    ]

    for label, count in summary["health_distribution"].items():
        summary_rows.append((label, count, "number"))

    summary_rows.extend([
        ("", "", "blank"),
        ("Top 10 Users by Engagement Depth", "", "section"),
    ])
    if summary["top_users"]:
        for row in summary["top_users"]:
            summary_rows.append((row["user_login"], f"depth: {row['engagement_depth']}", "label"))
    else:
        summary_rows.append(("None", "", "label"))

    summary_rows.extend([
        ("", "", "blank"),
        ("Users Needing Enablement", "", "section"),
        ("Count", len(summary["enablement_users"]), "number"),
        ("Details", "See 'Needs Enablement' sheet", "label"),
    ])

    current_row = 1
    for label, value, row_type in summary_rows:
        summary_ws.cell(row=current_row, column=1, value=label)
        summary_ws.cell(row=current_row, column=2, value=value)
        if row_type == "section":
            summary_ws.cell(row=current_row, column=1).font = SECTION_FONT
        elif row_type in {"label", "number", "pct"}:
            summary_ws.cell(row=current_row, column=1).font = LABEL_FONT
        if row_type == "number":
            summary_ws.cell(row=current_row, column=2).number_format = "#,##0"
        elif row_type == "pct":
            summary_ws.cell(row=current_row, column=2).number_format = "0.0"
        current_row += 1

    _autosize_columns(summary_ws)

    try:
        workbook.save(output_path)
        return output_path
    except PermissionError:
        timestamp = datetime.now().strftime("%H%M%S")
        retry_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        workbook.save(retry_path)
        return retry_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate a formatted Excel productivity report from GitHub Copilot metrics.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python copilot_productivity_report.py --enterprise my-ent
  python copilot_productivity_report.py --orgs org1,org2 --token ghp_xxx
  python copilot_productivity_report.py --enterprise my-ent --output-dir ./reports
        """,
    )
    parser.add_argument("--token", default=os.environ.get("GITHUB_TOKEN", ""))
    parser.add_argument("--enterprise", default=os.environ.get("ENTERPRISE_SLUG", ""))
    parser.add_argument("--orgs", default=os.environ.get("ORGS", ""))
    parser.add_argument("--output-dir", default=os.environ.get("OUTPUT_DIR", "."))
    parser.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help=(
            "Save raw API payloads to ./debug_*.{ndjson,json} and print the first "
            "record's field names per org. Useful for diagnosing field-name mismatches."
        ),
    )
    return parser.parse_args()


def main() -> None:
    if load_dotenv:
        env_path = Path(__file__).parent / ".env"
        if env_path.exists():
            load_dotenv(env_path)

    args = parse_args()
    token = args.token or os.environ.get("GITHUB_TOKEN", "")
    if not token:
        sys.exit("ERROR: No token. Use --token or set GITHUB_TOKEN.")

    print()
    validate_token(token)
    print()

    if args.debug:
        print("🐛 Debug mode ON — saving raw payloads to ./debug_*.{ndjson,json}\n")

    orgs_str = args.orgs or os.environ.get("ORGS", "")
    enterprise = args.enterprise or os.environ.get("ENTERPRISE_SLUG", "")

    if orgs_str:
        orgs = [org.strip() for org in orgs_str.split(",") if org.strip()]
    elif enterprise:
        print(f"🏢 Discovering orgs under: {enterprise}")
        orgs = discover_orgs(token, enterprise)
        if not orgs:
            sys.exit("ERROR: No orgs found.")
        print(f"   Found {len(orgs)} org(s)\n")
    else:
        sys.exit("ERROR: Use --enterprise or --orgs.")

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    output_path = output_dir / f"copilot_productivity_{today}.xlsx"

    print(f"📊 Copilot Productivity Report — {REPORT_DAYS}-day window\n")

    # Pre-fetch enterprise seats (if --enterprise provided). The seats response
    # carries a `seat.organization.login` per seat, which lets us:
    #   (1) Discover enterprise-owned orgs even when /enterprises/{ent}/organizations
    #       is forbidden — different endpoints have different access requirements.
    #   (2) Build a GLOBAL seat map so a user holding an enterprise-level seat
    #       granted by org A shows that seat in any org B's row (Copilot
    #       Enterprise spillover — the user's activity appears in B's NDJSON
    #       because they're a member, but the licensing seat lives in A).
    ent_seats_by_org: dict[str, list[dict]] = {}
    ent_status = ""
    if enterprise:
        print(f"🏢 Pre-fetching enterprise-level seats for: {enterprise}")
        ent_seats_by_org, ent_status = fetch_enterprise_seats(
            token, enterprise, debug=args.debug
        )
        if ent_seats_by_org:
            total_ent_seats = sum(len(v) for v in ent_seats_by_org.values())
            ent_org_names = sorted(
                k for k in ent_seats_by_org if k != "__enterprise_team__"
            )
            print(
                f"   ✅ Enterprise endpoint returned {total_ent_seats} seat(s) "
                f"across {len(ent_org_names)} org(s): {ent_org_names}"
            )
            # Discovery: add any enterprise-owned orgs we don't already know about.
            existing_lower = {o.lower() for o in orgs}
            for ent_org in ent_org_names:
                if ent_org.lower() not in existing_lower:
                    orgs.append(ent_org)
                    existing_lower.add(ent_org.lower())
                    print(f"   ➕ Added enterprise-owned org to iteration: {ent_org}")
        else:
            print(
                f"   ⚠ Enterprise seats endpoint returned no data (status: {ent_status}). "
                f"Falling back to per-org seats only."
            )
        print()

    org_payloads: list[dict[str, Any]] = []
    seat_status: dict[str, str] = {}
    report_starts: list[str] = []
    report_ends: list[str] = []
    all_records: list[dict] = []
    empty_seat_orgs_with_activity: list[str] = []

    # Build a case-folded lookup for enterprise seats by org name so we can
    # merge them into the per-org payload regardless of casing differences.
    ent_seats_by_org_lower = {k.lower(): v for k, v in ent_seats_by_org.items()}

    for org in orgs:
        print(f"🔍 {org}")
        seats, status = fetch_seats(token, org, debug=args.debug)
        # Merge in any enterprise-level seats GitHub recorded for this org.
        # Use case-insensitive matching because the enterprise endpoint may
        # return org names with different casing than what was iterated.
        ent_for_org = ent_seats_by_org_lower.get(org.lower(), [])
        if ent_for_org and not seats:
            seats = ent_for_org
            status = "ok_enterprise"
            print(
                f"   🔁 Backfilled {len(seats)} seat(s) from enterprise endpoint "
                f"for {org}."
            )
        seat_status[org] = status
        print(f"   Seats: {len(seats)} (status: {status})")

        ndjson_records, report_start, report_end = fetch_ndjson(token, org, debug=args.debug)
        print(f"   NDJSON records: {len(ndjson_records)}")

        if status == "ok" and not seats and ndjson_records:
            empty_seat_orgs_with_activity.append(org)

        if report_start:
            report_starts.append(report_start)
        if report_end:
            report_ends.append(report_end)

        if not seats and not ndjson_records:
            print("   ⚠ No data — skipping.")
            continue

        org_payloads.append({
            "org": org,
            "seats": seats,
            "records": ndjson_records,
            "report_end": report_end,
        })
        all_records.extend(ndjson_records)

    if not org_payloads:
        sys.exit("ERROR: No reportable user data found.")

    # Build the GLOBAL seat map (login.lower() → seat) across every seat we
    # collected (per-org + enterprise + enterprise-team-only seats). This is
    # the fallback consulted by build_user_rows for users that appear in an
    # org's NDJSON but don't have a seat granted by THAT org — typically
    # because their seat was granted by another org under the same enterprise
    # or by an enterprise team.
    global_seat_map: dict[str, dict] = {}
    for payload in org_payloads:
        for seat in payload["seats"]:
            login = ((seat.get("assignee") or {}).get("login") or "").lower()
            if login:
                global_seat_map[login] = seat
    # Also include enterprise-team-only seats (no `organization` field) that
    # wouldn't otherwise be reachable from any per-org payload.
    for seat in ent_seats_by_org.get("__enterprise_team__", []):
        login = ((seat.get("assignee") or {}).get("login") or "").lower()
        if login and login not in global_seat_map:
            global_seat_map[login] = seat
    if global_seat_map:
        print(
            f"\n🗂  Global seat map built: {len(global_seat_map)} licensed user(s) "
            f"across all sources. Will be used as a fallback for users whose "
            f"per-org seat lookup misses (Copilot Enterprise spillover)."
        )

    # Re-compute which orgs still have 0 seats + activity after the fallback.
    still_empty = [
        p["org"]
        for p in org_payloads
        if not p["seats"] and p["records"]
    ]
    if still_empty:
        print(
            f"\n⚠ {len(still_empty)} org(s) still have 0 seats but show Copilot activity in NDJSON:"
        )
        for org in still_empty:
            print(f"     • {org}")
        if global_seat_map:
            print(
                "\n   These users' seat data will be sourced from the global seat map "
                "(other orgs / enterprise endpoint) where available.\n"
            )
        else:
            print(
                "\n   What this means:\n"
                "   • The seats API is NOT failing — it's returning HTTP 200 with an empty list.\n"
                "   • Your PAT scopes/permissions are NOT the issue.\n"
                "   • Likely cause: the activity is from users on a personal Copilot Pro / Individual\n"
                "     subscription (not org-assigned seats), so there's no org-level seat record to read\n"
                "     a `created_at` from. `Seat Assigned Date` will be blank for these users —\n"
                "     that's accurate, not a bug.\n"
                "   • If the org IS supposed to have Copilot Business/Enterprise:\n"
                "       1) Pass --enterprise <slug> (NOT your username) to use the enterprise endpoint.\n"
                "       2) Confirm the org actually has assigned seats at\n"
                "          https://github.com/organizations/<org>/settings/copilot/seat_management.\n"
                "   • `Last Activity Date` is still backfilled from NDJSON, so the report remains useful.\n"
            )

    # Global dedupe — the user-level NDJSON returns each user's global metrics
    # for every org they hold a seat in. Without this, totals get multiplied by
    # the number of orgs each user belongs to.
    print(f"\n🔁 Deduping NDJSON across {len(org_payloads)} org(s) by (user_login, day) …")
    deduped_records, removed = dedupe_ndjson_records(all_records)
    if removed:
        print(
            f"   Collapsed {len(all_records)} records → {len(deduped_records)} "
            f"({removed} duplicate (user, day) pair(s) removed)."
        )
    else:
        print(f"   {len(all_records)} records, no cross-org duplicates found.")

    global_aggregation = aggregate_user_ndjson(deduped_records)
    print(f"   📊 Unique users with activity: {len(global_aggregation)}")

    overall_report_end = max(report_ends) if report_ends else ""

    all_user_rows: list[dict[str, Any]] = []
    for payload in org_payloads:
        user_rows = build_user_rows(
            payload["org"],
            payload["seats"],
            payload["records"],
            report_end=payload["report_end"] or overall_report_end,
            global_aggregation=global_aggregation,
            global_seat_map=global_seat_map,
            debug=args.debug,
        )
        if args.debug:
            with_dates = sum(1 for r in user_rows if r.get("seat_assigned_date"))
            print(
                f"   📊 {payload['org']}: {len(user_rows)} row(s) "
                f"({with_dates} with seat_assigned_date populated)"
            )
        else:
            print(f"   📊 {payload['org']}: {len(user_rows)} row(s)")
        all_user_rows.extend(user_rows)

    if not all_user_rows:
        sys.exit("ERROR: No reportable user data found.")

    all_user_rows.sort(key=lambda row: (row["organization"], row["user_login"].lower()))
    unique_rows = dedupe_users_across_orgs(
        all_user_rows,
        report_end=overall_report_end,
        metrics_are_global=True,
    )
    summary = build_team_summary(
        all_user_rows,
        min(report_starts) if report_starts else "",
        overall_report_end,
        unique_rows=unique_rows,
    )

    failed_orgs = [org for org, status in seat_status.items() if status != "ok"]
    if failed_orgs:
        print(
            f"\n⚠ Seat data was unavailable for {len(failed_orgs)} org(s): "
            f"{', '.join(failed_orgs)}.\n"
            f"   → 'Seat Assigned Date' and 'Last Activity Date' will be blank for users in those orgs.\n"
            f"   → Users who are licensed but INACTIVE will be MISSING from the report entirely\n"
            f"     (only NDJSON-active users show up), so Team Adoption Rate and Needs Enablement\n"
            f"     counts for those orgs are unreliable / likely overstated.\n"
            f"   → Ensure your token has 'manage_billing:copilot' scope AND you are an admin of those orgs."
        )

    if len(set(report_starts)) > 1 or len(set(report_ends)) > 1:
        print(
            f"\n⚠ Report windows differ across orgs:\n"
            f"   start days: {sorted(set(report_starts))}\n"
            f"   end days:   {sorted(set(report_ends))}\n"
            f"   → A user's global activity may span up to the widest window; the canonical\n"
            f"     28-day window may be exceeded for some users."
        )

    print("\n📊 Writing Excel report …")
    final_path = write_excel(output_path, all_user_rows, summary)
    print(f"✅ Done! Report saved to {final_path.resolve()}\n")


if __name__ == "__main__":
    main()
